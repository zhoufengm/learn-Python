
import os
import openpyxl
from openpyxl.utils import get_column_letter,column_index_from_string
import openpyxl.styles  #字体样式
from openpyxl import load_workbook
import wsproto

'''
path = r"E:\AutoCode\learn-Python"
# 修改工作路径
os.chdir(path)

workbook = openpyxl.load_workbook('test.xlsx')
# 打印Excel表中的所有表
# print(workbook.sheetnames)
# >>> ['测试用例集', '启动APP', '个人信息', '退出APP', '测试结果统计', '测试结果明细']

# 获取指定sheet表
# sheet = workbook['启动APP']
sheet = workbook['测试用例集']
# 获取活动表
# 活动表：上一次关闭时所在的工作表
# sheet = workbook.active
print('当前工作表是：',sheet)
# >>>  <Worksheet "启动APP">
# 疑问：为什么不是第一个?

  # 获取表格的尺寸大小
# print(sheet.dimensions)   # 没有这个方法了？


# 获取单元格中的数据
# 方法1：指定坐标的方式
cell1 = sheet['A1'].value     # 获取A1单元格的数据
cell2 = sheet['C3'].value     # 获取C3单元格的数据
print('cell1 : ',cell1)
print('cell2 : ',cell2)

# 方法2： 指定行列的方式
cell3 = sheet.cell(row=1,column=1).value
cell4 = sheet.cell(row=3,column=4).value
print('cell3 : ',cell3)     # 获取第1行第1列的数据
print('cell4 : ',cell4)      # 获取第3行第4的数据

# 4.2：获取单元格的行、列、坐标
# cell5 = sheet.cell(row=2,column=1).value
# print('cell5 : ',cell5)
# print(cell5,cell5.row,cell5.column,cell5.coordinate)
# b报错了：AttributeError: 'str' object has no attribute 'row'

# 5：获取区间内的数据
# 获取单行单列数据的时候，使用一层for循环；获取多行多列、指定区间的数据时，使用两层for循环

# 5.1 获取指定区间的数据
cell6 = sheet['A1:A5']   # 获取A1到A5的数据
print(cell6)
# >>> ((<Cell '测试用例集'.A1>,), (<Cell '测试用例集'.A2>,), (<Cell '测试用例集'.A3>,), (<Cell '测试用例集'.A4>,), (<Cell '测试用例集'.A5>,))

# 打印A1到A5的数据
print('打印A1到A5的数据')
for i in cell6:
    for j in i:
        print(j.value)
# 用例名称
# 启动APP
# 设置个人信息
# 退出APP
# None

# 5.2 获取指定行列的数据
# sheet[“A”] — 获取A列的数据
# sheet[“A:C”] — 获取A,B,C三列的数据
# sheet[5] — 只获取第5行的数据
cell7 = sheet['2']   # 获取第2行的数据
print('打印第二行的数据：')
for i in cell7:
    print(i.value)

cell8 = sheet['A:B']    # 获取AB列的数据
print('打印AB列数据: ')
for i in cell8:
    for j in i:
        print(j.value)


# 5.3 按行、列获取值
# iter_rows()：按行读取
# iter_cols()：按列读取
print('按行获取值：')
for i in sheet.iter_rows(min_row=2,max_row=4,min_col=1,max_col=2):
    for j in i:
        # 获取第2行到第4行，前两列数据, 按行一行一行打印
        print(j.value)

print('按列获取值： ')
for i in sheet.iter_cols(min_row=2,max_row=4,min_col=1,max_col=2):
    for j in i:
        # 获取第1列和第2列中 第2行到第4行的数据，按列一列一列打印
        print(j.value)


# 5.4 获取活动表的行列数
# 方法1：使用
# sheet.max_row 获取行数
# sheet.max_column 获取列数
row1 = sheet.max_row
column1 = sheet.max_column
print('行数rows1: ', row1)
print('列数column1: ', column1)

# 方法2：自己写一个for循环
row2 = []
column2 = []
# 获取当前活动表有多少行
for i in sheet.rows:
    row2.append(list(i))    # i是元组类型，转为列表

# 获取当前活动表有多少列
for i in sheet.columns:
    column2.append(list(i))    # i是元组类型，转为列表

print('行数rows2: ', str(len(row2)))
print('列数column2: ', str(len(column2)))




# 6 创建新的excel
path2 = r"E:\AutoCode\learn-Python\learn-appium"
os.chdir(path2)

workbook = openpyxl.load_workbook('test1.xlsx')
sheet = workbook.active
print('当前活动表名：' +str(sheet))



sheet.title = '车辆信息' # 这个操作相当于修改sheet表名称：将活动表名称修改为指定名字。
# workbook.save('1.xlsx') 
# 这个操作相当于保存：保存时如果使用原来的名字，就直接保存；
# 如果使用了别的名字，就会另存为一个新文件，相当于另存为
# 上一步的修改名称也体现在新文件中，原来的文件test1没有修改
# ！！！！执行保存操作时该文件不能打开，否则会报错

# 修改单元格数据的两种方式
sheet['A1'] = '车辆编号'
sheet['B1'].value = '车辆类型'
workbook.save('test1.xlsx')



# 7 添加数据
# 使用append()方法，在原来数据的后面，按行插入数据
data = [
    [3,805,'嵩屿码头','天竺山',72]  # 数字默认靠右对齐，文本默认靠左对齐
    ['4','850','嵩屿码头','文化宫','72']    # 加了引号，都靠左对齐
]
for row in data:
    sheet.append(row)
workbook.save('test1.xlsx')


# 插入空行空列
# insert_rows(idx=数字编号, amount=要插入的行数)，插入的行数是在idx行数的上方插入
# insert_cols(idx=数字编号, amount=要插入的列数)，插入的位置是在idx列数的左侧插入
sheet.insert_rows(idx=3,amount=2)
sheet.insert_cols(idx=5,amount=1)
workbook.save('test1.xlsx')


# 删除行、列
# delete_rows(idx=数字编号, amount=要删除的行数)
# delete_cols(idx=数字编号, amount=要删除的列数)
# sheet.delete_rows(idx=3)   # 删除第3行
# sheet.delete_cols(idx=5)   # 删除第5列
sheet.delete_rows(idx=3,amount=2)   # 删除第3行，及往下共2行
sheet.delete_cols(idx=5,amount=2)   # 删除第5列，及往右共2列
workbook.save('test1.xlsx')


# 8 移动指定区间的单元格
# move_range(“数据区域”,rows=,cols=)：正整数为向下或向右、负整数为向左或向上
sheet.move_range('C4:D5',rows=3,cols=2) # 移动C4:D5构成的矩形格子(原区域数据就没了)
workbook.save('test1.xlsx')


# 9 字母列号与数字列号之间的转换
# 根据列的数字返回字母
print(get_column_letter(2)) # B
# 根据字母返回列的数字
print(column_index_from_string('D')) #4


# 10 字体样式
# openpyxl.styles.Font(name=字体名称,size=字体大小,bold=是否加粗,italic=是否斜体,color=字体颜色)
# 字体颜色中的color是RGB的16进制表示
# 10.1 查看字体样式
cellz = sheet['A1']
font = cellz.font
print('当前单元格 %s',cellz,' 的字体样式是：')
print(font.name,font.size,font.bold,font.italic,font.color)
# 当前单元格 %s 的字体样式是： <Cell '车辆信息'.A1>
# 等线 11.0 False False <openpyxl.styles.colors.Color object>
# Parameters:
# rgb=None, indexed=None, auto=None, theme=1, tint=0.0, type='theme'

#  10.2 修改字体样式
cellz.font = openpyxl.styles.Font(name='微软雅黑',size=20,bold=True,italic=True,color='FF0000')
workbook.save('test1.xlsx')


# 10.3 可以使用for循环，修改多行多列的数据
# cellt = sheet['A']  #修改第一列数据
cellt = sheet['1']  #修改第一行数据
for i in cellt:
    i.font =openpyxl.styles.Font(name='黑体',size=15,bold=True,italic=False,color='FF0000')
workbook.save('test1.xlsx')


# 10.4 设置对齐格式
# Alignment(horizontal=水平对齐模式,vertical=垂直对齐模式,text_rotation=旋转角度,wrap_text=是否自动换行)
# 水平对齐：‘distributed’，‘justify’，‘center’，‘left’， ‘centerContinuous’，'right，‘general’
# 垂直对齐：‘bottom’，‘distributed’，‘justify’，‘center’，‘top’
cell4 = sheet['A1']
alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',text_rotation=0,wrap_text=False)
cell4.alignment = alignment
workbook.save('test1.xlsx')


cell4_1 = sheet['A']
alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',text_rotation=0,wrap_text=False)
for i in cell4_1:
    i.alignment = alignment
workbook.save('test1.xlsx')


# 10.5 设置行高列宽
# row_dimensions[行编号].height = 行高
# column_dimensions[列编号].width = 列宽
sheet.row_dimensions[1].height = 50
sheet.column_dimensions['B'].width = 20
workbook.save('test1.xlsx')



path3 = r"E:\AutoCode\learn-Python"
os.chdir(path3)
workbook = openpyxl.load_workbook('test.xlsx')
sheet = workbook.active
print('当前活动表名：' +str(sheet))


# 设置所有单元格
width = 23.0 #设置宽度
height = 32.0  #设置高度
print(height)
# 打印行数和列数
print("row:",sheet.max_row,"column；",sheet.max_column)

for i in range(1,sheet.max_row+1):
    sheet.row_dimensions[i].height = height
for i in range (1,sheet.max_column+1):
    sheet.column_dimensions[get_column_letter(i)].width = width
workbook.save('test.xlsx')


# 11 合并、拆分单元格
# merge_cells(start_row=起始行号，start_column=起始列号，end_row=结束行号，end_column=结束列号)
# unmerge_cells(start_row=起始行号，start_column=起始列号，end_row=结束行号，end_column=结束列号)
# 方法1
sheet.merge_cells('A5:B7')
# 方法2
sheet.merge_cells(start_row=7,start_column=3,end_row=9,end_column=4)
workbook.save('test.xlsx')


# 加一个居中对齐
cell11 = sheet['A1']
alignment =  openpyxl.styles.Alignment(horizontal='center',vertical='center',text_rotation=0,wrap_text=0)
cell11.alignment = alignment

cell12 = sheet['C2']
alignment =  openpyxl.styles.Alignment(horizontal='center',vertical='center',text_rotation=0,wrap_text=0)
cell12.alignment = alignment
workbook.save('test.xlsx')


# 12 创建新的sheet(create_sheet)
path4 = r"E:\AutoCode\learn-Python"
os.chdir(path4)
workbook = openpyxl.load_workbook('test.xlsx')
# sheet = workbook['3号sheet']
# print('当前活动表名：' +str(sheet))

# workbook.create_sheet('3号sheet')   #创建新的sheet表
# print(workbook.sheetnames)

# 12.1 修改sheet名字(title)
# sheet.title = '2号sheet'

# 复制sheet表(copy_worksheet)
sheet = workbook['2号sheet']
# workbook.copy_worksheet(sheet)

# 删除sheet表(remove)
workbook.remove(sheet)
print(workbook.sheetnames)
workbook.save('test.xlsx')

'''
# 13 操作多个Excel表


