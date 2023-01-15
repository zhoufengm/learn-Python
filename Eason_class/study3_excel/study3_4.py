# 2023.1.13 数据匹配和筛选
from openpyxl import load_workbook,Workbook

'''
# 案例四
wb = load_workbook('./study3_excel/Excel实操文件/10月员工绩效表.xlsx')
ws = wb.active

# 创建员工信息字典
staff_dict = {}
# 获取ws中除表头之外的信息
for row in ws.iter_rows(min_row=2,values_only=True):
    staff_id = row[0]
    # 添加字典信息
    staff_dict[staff_id] = row

# 提示输入员工工号
id = input("请输入员工工号：")
if id in staff_dict.keys():
    print("该员工信息如下：")
    print(staff_dict[id])
else:
    print("找不到该员工信息。")

'''

# 数据筛选
wb = load_workbook('./study3_excel/Excel实操文件/10月考勤统计.xlsx')
ws = wb.active

'''
# 新建工作簿保存迟到人员信息
new_wb = Workbook()
new_ws = new_wb.active

# 设置表头
new_ws.append(['工号','姓名','部门','迟到时间(分钟)','迟到次数(次数)'])
# 从第二行开始遍历表格
for row in ws.iter_rows(min_row=2,values_only=True):
    # 时间大于45分钟且次数大于3次为迟到
    if row[3] > 45 and row[4] > 3:
        print(row)
        new_ws.append(row)

# 保存迟到人员信息工作簿
new_wb.save('./study3_excel/10月迟到人员信息.xlsx')
'''
# 创建字典存储员工迟到次数
info_dict = {}
for row in ws.iter_rows(min_row=2,values_only=True):
    # 获取员工id
    staff_id = row[0]
    # 获取最后一列迟到次数
    staff_late = row[-1]
    info_dict[staff_id] = staff_late

# print(info_dict)

# 打开手动统计的迟到次数工作簿
a_wb = load_workbook('./study3_excel/Excel实操文件/迟到次数月度统计（10月更新）.xlsx')
a_ws = a_wb.active

# 遍历第三行~最后一行，第1列到第13列
for row in a_ws.iter_rows(min_row=3, max_col=13, values_only=True):
    late_id = row[0]
    late_count = row[-1]
    # 使用当然工号去字典里查询对比
    if late_count != info_dict[late_id]:
        print(f'工号：{late_id}迟到次数不匹配，请检查！')