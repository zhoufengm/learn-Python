# 2023.1.13
from openpyxl import load_workbook,Workbook

# 案例三
# 打开原工作薄
odd_wb = load_workbook('E:/AutoCode/Eason_class/study3_excel/Excel实操文件/10月员工绩效表.xlsx')
odd_ws = odd_wb.active

# 新建一个工作簿
new_wb = Workbook()
new_ws = new_wb.active

# 设置表头
headers = ['工号','姓名','部门','基本工资','绩效','提成','应付工资','是否确认']
new_ws.append(headers)

# 从第二行开始更新信息
for row in odd_ws.iter_rows(min_row =2,values_only =True):
    # row[0]这个取索引要减1
    pay = row[3] + row[4] + row[5]
    staff = [row[0], row[1], row[2], row[5], row[3], row[4], pay, row[6]]
    new_ws.append(staff)

new_wb.save('E:/AutoCode/Eason_class/study3_excel/10月员工工资表.xlsx')