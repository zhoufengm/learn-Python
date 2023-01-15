from openpyxl import load_workbook,Workbook

# 课后练习
staff_wb = load_workbook('./practice.xlsx')
active_ws = staff_wb['上半年公司名单']
for cell in active_ws.iter_rows(min_col=2,max_col=3,min_row=8,max_row=10,values_only=True):
    print(cell)


'''
# 课堂练习1
# 使用load_workbook打开工作簿
workbook1 = load_workbook('./practice.xlsx')
print(workbook1)

# 使用Workbook新建工作簿对象
workbook2 = Workbook()
print(workbook2)
# 将新建的工作簿保存为practice1.xlsx
workbook2.save('practice1.xlsx')
'''

'''
# 课堂练习2
workbook3 =  load_workbook('./practice.xlsx')

# 使用active属性获取活动工作表
active_ws = workbook3.active
print(active_ws)

# 按表名取表
shy = workbook3['上半年公司名单']
xhy = workbook3['下半年公司名单']

# 打印
print(workbook3)
print(shy)
print(xhy)

'''

'''
# 课堂练习3
# 打开工作簿
staff_wb = load_workbook('./practice.xlsx')

# 获取活动工作表
active_ws = staff_wb.active

# 使用append 添加2行数据 ,元组，列表都可以
info_tuple = ('S1912','王多鱼',5000,'销售')
info_list = ['S1913','周大壮',6000,'工程师']
active_ws.append(info_tuple)
active_ws.append(info_list)

# 另存为
staff_wb.save('practice1.xlsx')

# 获取活动工作表（之前保存关闭了，需要重新获取，否则后面打印内容不对）
active_ws = staff_wb.active

# 获取工作表所有数据
# values_only=True 打印出来是值，一般读取时用
# ('S1021', '范桂香', 6000, '后勤部')
data_all = active_ws.iter_rows(values_only=True)

# values_only=False 打印出来是对象，一般写入时用
# (<Cell '上半年公司名单'.A23>, <Cell '上半年公司名单'.B23>, <Cell '上半年公司名单'.C23>, <Cell '上半年公司名单'.D23>)
# data_all = active_ws.iter_rows(values_only=False)

# 遍历打印所有数据
for staff in data_all:
    print(staff)

# # 获取指定行列数据
# data_some = active_ws.iter_rows(min_row=5,max_row=7,min_col=1,max_col=3,values_only=True)
# for row in data_some:
#     print(row)
#     # print(row[1])  # 打印第2列数据
#     # 若想打印这行的每一个单元格，可以再加一层循环
#     for cell in row：
#           print(cell)

# # []获取单行或单列，输出<Cell'工作表名称'.坐标>形式的元组
# print(active_ws['A'])
# print(active_ws[3])
# print(active_ws['B4'])
'''


'''
# 课堂练习4
# 打开工作簿
staff_wb = load_workbook('./practice.xlsx')

# 获取指定工作表
staff_ws = staff_wb['下半年公司名单']

# 循环获取第四列（D列)所有单元格对象
for cell in staff_ws['D']:
    # 如果为表头，则跳过本次循环
    if cell.value == '部门':
        continue
    # 打印原有的值
    print("原来的部门：" + str(cell.value))
    # 将原有的值修改为"战略储备部"
    cell.value = '战略储备部'

# 将结果另存为
staff_wb.save('./practice1.xlsx')
'''